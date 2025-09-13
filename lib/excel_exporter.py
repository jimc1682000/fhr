"""Excel export helpers for attendance analyzer."""

from typing import List, Tuple, Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

from attendance_analyzer import Issue, IssueType


def init_workbook() -> Tuple[Workbook, Any, Font, PatternFill, Border, Alignment]:
    """Initialize workbook, worksheet, and basic styles."""
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤分析"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    return wb, ws, header_font, header_fill, border, center_alignment


def write_headers(
    ws,
    headers: List[str],
    header_font: Font,
    header_fill: PatternFill,
    border: Border,
    alignment: Alignment,
) -> None:
    """Write header row with styles."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border


def write_status_row(
    ws,
    last_date: str,
    complete_days: int,
    last_analysis_time: str,
    border: Border,
    alignment: Alignment,
) -> int:
    """Write incremental status row and return next data row."""
    ws.cell(row=2, column=1).value = last_date
    ws.cell(row=2, column=2).value = "狀態資訊"
    ws.cell(row=2, column=3).value = 0
    ws.cell(row=2, column=4).value = (
        f"📊 增量分析完成，已處理至 {last_date}，共 {complete_days} 個完整工作日 | 上次分析時間: {last_analysis_time}"
    )
    ws.cell(row=2, column=5).value = ""
    ws.cell(row=2, column=6).value = "上次處理範圍內無新問題需要申請"
    ws.cell(row=2, column=7).value = "系統狀態"
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for col in range(1, 8):
        cell = ws.cell(row=2, column=col)
        cell.fill = gray_fill
        cell.border = border
        if col in [1, 2, 3, 5, 7]:
            cell.alignment = alignment
    return 3


def write_issue_rows(
    ws,
    issues: List[Issue],
    start_row: int,
    incremental_mode: bool,
    border: Border,
    alignment: Alignment,
) -> None:
    """Write issue rows into the worksheet."""
    for row_idx, issue in enumerate(issues, start_row):
        date_cell = ws.cell(row=row_idx, column=1)
        date_cell.value = issue.date.strftime("%Y/%m/%d")
        date_cell.alignment = alignment
        date_cell.border = border

        type_cell = ws.cell(row=row_idx, column=2)
        type_cell.value = issue.type.value
        type_cell.alignment = alignment
        type_cell.border = border
        if issue.type == IssueType.LATE:
            type_cell.fill = PatternFill(
                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
            )
        elif issue.type == IssueType.OVERTIME:
            type_cell.fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
            )
        elif issue.type == IssueType.WFH:
            type_cell.fill = PatternFill(
                start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
            )
        elif issue.type == IssueType.FORGET_PUNCH:
            type_cell.fill = PatternFill(
                start_color="FFF0E6", end_color="FFF0E6", fill_type="solid"
            )

        duration_cell = ws.cell(row=row_idx, column=3)
        duration_cell.value = issue.duration_minutes
        duration_cell.alignment = alignment
        duration_cell.border = border

        desc_cell = ws.cell(row=row_idx, column=4)
        desc_cell.value = issue.description
        desc_cell.border = border

        range_cell = ws.cell(row=row_idx, column=5)
        range_cell.value = issue.time_range
        range_cell.alignment = alignment
        range_cell.border = border

        calc_cell = ws.cell(row=row_idx, column=6)
        calc_cell.value = issue.calculation
        calc_cell.border = border

        if incremental_mode:
            status_cell = ws.cell(row=row_idx, column=7)
            status_cell.value = "[NEW] 本次新發現" if issue.is_new else "已存在"
            status_cell.alignment = alignment
            status_cell.border = border
            if issue.is_new:
                status_cell.fill = PatternFill(
                    start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                )


def set_column_widths(ws, incremental_mode: bool) -> None:
    """Set worksheet column widths.
    - Base width 15 for most columns
    - Description (D) wider for readability
    - Calculation (F) widest; wider when incremental_mode is on
    - Status (G) visible without wrapping when incremental_mode is on
    """
    col_count = 7 if incremental_mode else 6
    for col in range(1, col_count + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["F"].width = 40 if incremental_mode else 35
    if incremental_mode:
        ws.column_dimensions["G"].width = 24


def save_workbook(wb: Workbook, filepath: str) -> None:
    """Persist workbook to disk."""
    wb.save(filepath)
