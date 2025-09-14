import os
import tempfile
import unittest

from openpyxl import load_workbook

from attendance_analyzer import AttendanceAnalyzer


class TestExportExcelIncrementalIssueRow(unittest.TestCase):
    def test_issue_row_includes_status_when_incremental(self):
        # Build a minimal file that will produce at least one issue (late)
        text = (
            '應刷卡時段\t當日卡鐘資料\t刷卡別\t卡鐘編號\t資料來源\t異常狀態\t處理狀態\t異常處理作業\t備註\n'
            '2025/07/01 08:30\t2025/07/01 12:45\t上班\t1\t刷卡匯入\t\t\t\t\n'
            '2025/07/01 17:30\t2025/07/01 18:00\t下班\t1\t刷卡匯入\t\t\t\t\n'
        )
        with tempfile.TemporaryDirectory() as tmp:
            src = os.path.join(tmp, '202507-測試人-出勤資料.txt')
            with open(src, 'w', encoding='utf-8') as f:
                f.write(text)

            an = AttendanceAnalyzer()
            # Parse in full mode to avoid skipping due to persisted state
            an.parse_attendance_file(src, incremental=False)
            an.group_records_by_day()
            an.analyze_attendance()

            # Manually enable incremental flag for export branch coverage
            self.assertGreater(len(an.issues), 0)
            an.incremental_mode = True

            out_xlsx = os.path.join(tmp, 'out.xlsx')
            an.export_excel(out_xlsx)

            wb = load_workbook(out_xlsx)
            ws = wb.active
            # Header should include status column
            self.assertEqual(ws['G1'].value, '狀態')
            # First data row should be the issue, with status value present
            self.assertIn(ws['G2'].value, ('[NEW] 本次新發現', '已存在'))


if __name__ == '__main__':
    unittest.main()
"""Category: Export/Excel
Purpose: Ensure incremental export writes status column on issue rows."""
