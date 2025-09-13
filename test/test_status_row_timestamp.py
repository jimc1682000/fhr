import os
import tempfile
import csv
import unittest
from openpyxl import load_workbook

from attendance_analyzer import AttendanceAnalyzer


class TestStatusRowTimestamp(unittest.TestCase):
    def _run_clean_case(self):
        text = """應刷卡時段	當日卡鐘資料	刷卡別	卡鐘編號	資料來源	異常狀態	處理狀態	異常處理作業	備註
2025/07/01 08:00	2025/07/01 09:00	上班	1	刷卡匯入				
2025/07/01 17:00	2025/07/01 18:00	下班	1	刷卡匯入				
"""
        tmpdir = tempfile.mkdtemp()
        valid_named = os.path.join(tmpdir, "202507-王小明-出勤資料.txt")
        with open(valid_named, "w", encoding="utf-8") as f:
            f.write(text)
        analyzer = AttendanceAnalyzer()
        analyzer.incremental_mode = True
        analyzer.parse_attendance_file(valid_named)
        analyzer.group_records_by_day()
        analyzer.analyze_attendance()
        return analyzer, valid_named

    def test_csv_status_has_timestamp(self):
        analyzer, named_path = self._run_clean_case()
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            out_csv = f.name
        try:
            analyzer.export_csv(out_csv)
            with open(out_csv, encoding="utf-8-sig") as f:
                rows = list(csv.reader(f, delimiter=";"))
            self.assertGreaterEqual(len(rows), 2)
            status_line = rows[1]
            self.assertIn("上次分析時間:", status_line[3])
        finally:
            if os.path.exists(out_csv):
                os.unlink(out_csv)
            if os.path.exists(named_path):
                os.unlink(named_path)
            try:
                os.rmdir(os.path.dirname(named_path))
            except OSError:
                pass

    def test_excel_status_has_timestamp(self):
        analyzer, named_path = self._run_clean_case()
        with tempfile.NamedTemporaryFile(mode="w", suffix=".xlsx", delete=False) as f:
            out_xlsx = f.name
        try:
            analyzer.export_excel(out_xlsx)
            wb = load_workbook(out_xlsx)
            ws = wb.active
            self.assertIn("上次分析時間:", ws["D2"].value)
        finally:
            if os.path.exists(out_xlsx):
                os.unlink(out_xlsx)
            if os.path.exists(named_path):
                os.unlink(named_path)
            try:
                os.rmdir(os.path.dirname(named_path))
            except OSError:
                pass


if __name__ == "__main__":
    unittest.main()
"""Category: Export/StatusRow
Purpose: Status row includes last analysis timestamp for CSV/Excel."""
